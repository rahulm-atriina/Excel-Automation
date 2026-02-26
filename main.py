import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.title("Dual Source Agreement Transposition Engine")

# =====================================================
# CONFIG
# =====================================================

FINAL_METRIC_ORDER = [
    "GSR (Q)",
    "Target Quantity (Q)",
    "Target Value (Q)",
    "Invoiced Quantity(Q)",
    "Achievement % (Q)",
    "GSR (B)",
    "Target Quantity (B)",
    "Target Value (B)",
    "Invoiced Quantity(B)",
    "Achievement % (B)",
    "Rebate Rate",
    "Clear Cut CN"
]

BASE_COLUMNS = ["Customer", "Region", "Depot Code"]

# =====================================================
# OPTIMIZED TRANSFORMATION ENGINE
# =====================================================

@st.cache_data(show_spinner=False)
def transform_fast(df, block_column, block_metric_map, prefix=None):

    df = df.copy()
    df.columns = df.columns.str.strip()

    for col in BASE_COLUMNS:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Convert block column to category (BIG SPEED BOOST)
    df[block_column] = df[block_column].astype("category")

    # Only process selected metrics (performance gain)
    selected_metrics = list(set(sum(block_metric_map.values(), [])))
    available_metrics = [
        m for m in selected_metrics if m in df.columns
    ]

    if not available_metrics:
        return pd.DataFrame()

    grouped = (
        df.groupby(BASE_COLUMNS + [block_column], observed=True)[available_metrics]
        .sum()
        .reset_index()
    )

    if grouped.empty:
        return pd.DataFrame()

    # Faster than pivot()
    pivot_df = (
        grouped
        .set_index(BASE_COLUMNS + [block_column])[available_metrics]
        .unstack(block_column)
    )

    # Reorder to (block, metric)
    pivot_df = pivot_df.swaplevel(0, 1, axis=1)
    pivot_df.sort_index(axis=1, level=0, sort_remaining=False, inplace=True)

    selected_columns = []
    final_column_tuples = []

    for block in block_metric_map.keys():

        if block not in pivot_df.columns.levels[0]:
            continue

        display_block = f"{prefix}{block}" if prefix else block
        selected = block_metric_map.get(block, [])

        for metric in FINAL_METRIC_ORDER:
            if (
                metric in selected
                and (block, metric) in pivot_df.columns
            ):
                selected_columns.append((block, metric))
                final_column_tuples.append((display_block, "", metric))

    if not selected_columns:
        return pd.DataFrame()

    result = pivot_df.loc[:, selected_columns]
    result.columns = pd.MultiIndex.from_tuples(final_column_tuples)
    result = result.reset_index()

    return result


# =====================================================
# FAST EXCEL EXPORT (NO FULL CELL LOOPS)
# =====================================================

def export_with_borders(df):

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        # MUST keep index=True for MultiIndex columns
        df.to_excel(writer, index=True)

        sheet = writer.sheets["Sheet1"]

        thin = Side(style="thin")
        thick = Side(style="medium")

        max_row = sheet.max_row
        max_col = sheet.max_column

        # Hide index column visually (clean look)
        sheet.column_dimensions["A"].width = 2

        # Bold header rows
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.font = Font(bold=True)

        # Apply vertical thick borders between blocks
        col_map = {}

        # Start from column 2 because column 1 = index
        for col_index, col in enumerate(df.columns, start=2):

            if isinstance(col, tuple):
                block = col[0]
                if block not in BASE_COLUMNS:
                    col_map.setdefault(block, []).append(col_index)

        for cols in col_map.values():
            start = cols[0]
            end = cols[-1]

            for row in range(1, max_row + 1):
                sheet.cell(row=row, column=start).border = Border(left=thick)
                sheet.cell(row=row, column=end).border = Border(right=thick)

        # Auto column width
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

            sheet.column_dimensions[col_letter].width = min(max_len + 2, 35)

    buffer.seek(0)
    return buffer

# =====================================================
# UI
# =====================================================

main_file = st.file_uploader("Upload Main Raw File", type=["xlsx"])
eb_file = st.file_uploader("Upload EB Raw File", type=["xlsx"])

if main_file and eb_file:

    with st.spinner("Loading files..."):
        main_df = pd.read_excel(main_file, engine="openpyxl")
        eb_df = pd.read_excel(eb_file, engine="openpyxl")

    main_df.columns = main_df.columns.str.strip()
    eb_df.columns = eb_df.columns.str.strip()

    if "Component description" not in main_df.columns:
        st.error("Component description column missing in Main file.")
        st.stop()

    if "Deal Suffix" not in eb_df.columns:
        st.error("Deal Suffix column missing in EB file.")
        st.stop()

    # ================= MAIN METRIC GRID =================

    st.subheader("Main File - Metric Selection")

    main_blocks = main_df["Component description"].dropna().unique().tolist()

    main_metric_df = pd.DataFrame(
        True,
        index=main_blocks,
        columns=[m for m in FINAL_METRIC_ORDER if m in main_df.columns]
    )

    edited_main = st.data_editor(
        main_metric_df,
        num_rows="fixed",
        width="stretch"
    )

    main_metric_map = {
        comp: edited_main.columns[edited_main.loc[comp]].tolist()
        for comp in edited_main.index
    }

    st.markdown("### Main Component Order")

    ordered_main = st.multiselect(
        "Reorder Main Components (Top = First in Output)",
        options=main_blocks,
        default=main_blocks
    )

    if not ordered_main:
        ordered_main = main_blocks

    # ================= EB METRIC GRID =================

    st.subheader("EB File - Metric Selection")

    eb_blocks = eb_df["Deal Suffix"].dropna().unique().tolist()
    eb_display = [f"EB{b}" for b in eb_blocks]

    eb_metric_df = pd.DataFrame(
        True,
        index=eb_display,
        columns=[m for m in FINAL_METRIC_ORDER if m in eb_df.columns]
    )

    edited_eb = st.data_editor(
        eb_metric_df,
        num_rows="fixed",
        width="stretch"
    )

    eb_metric_map = {
        eb_blocks[i]: edited_eb.columns[
            edited_eb.loc[f"EB{eb_blocks[i]}"]
        ].tolist()
        for i in range(len(eb_blocks))
    }

    st.markdown("### EB Component Order")

    selected_eb_display = st.multiselect(
        "Reorder EB Deal Suffix (Top = First in Output)",
        options=eb_display,
        default=eb_display
    )

    if not selected_eb_display:
        selected_eb_display = eb_display

    ordered_eb = [
        eb_blocks[eb_display.index(label)]
        for label in selected_eb_display
    ]

    # ================= GENERATE =================

    if st.button("Generate Final Output"):

        with st.spinner("Processing..."):

            main_result = transform_fast(
                main_df,
                "Component description",
                {k: main_metric_map[k] for k in ordered_main}
            )

            eb_result = transform_fast(
                eb_df,
                "Deal Suffix",
                {k: eb_metric_map[k] for k in ordered_eb},
                prefix="EB"
            )

            if main_result.empty:
                final_result = eb_result
            elif eb_result.empty:
                final_result = main_result
            else:
                final_result = pd.merge(
                    main_result,
                    eb_result,
                    on=BASE_COLUMNS,
                    how="outer"
                )

        st.dataframe(final_result, width="stretch")

        styled_buffer = export_with_borders(final_result)

        st.download_button(
            label="Download Final Output",
            data=styled_buffer,
            file_name="merged_transposed_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )